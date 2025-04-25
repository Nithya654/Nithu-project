from openpyxl import load_workbook

wb = load_workbook("C:/Users/nithy/OneDrive/Documents/training 19 march/march25/dataset.xlsx")
#sheet name
sheet = wb["Sheet"] 

# Read a cell
print(sheet["A1"].value)

# Write to a cell
sheet["B2"] = "Hello from Hexaware!"

# Save changes
wb.save("sample_modified.xlsx")

#2========
import openpyxl
path=("C:/Users/nithy/OneDrive/Documents/training 19 march/march25/dataset.xlsx")
# To open the workbook 
wb_obj = openpyxl.load_workbook(path)
# Get workbook active sheet object
sheet_obj = wb_obj.active
# Cell object is created by using sheet object's cell() method.
cell_obj = sheet_obj.cell(row = 1, column = 1)
print(cell_obj.value)

# import openpyxl module
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row = 1, column = 1)
c1.value = "Hello"
c2 = sheet.cell(row= 1 , column = 2)
c2.value = "World"
c3 = sheet['A2']
c3.value = "Welcome"
c4 = sheet['B2']
c4.value = "Everyone"
wb.save("sample.xlsx")

# import openpyxl module 
import openpyxl
wb = openpyxl.load_workbook("sample.xlsx") 
sheet = wb.active
c = sheet['A3'] 
c.value = "New Data"
wb.save("sample.xlsx")
